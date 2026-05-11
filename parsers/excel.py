from pathlib import Path


EXPECTED_SHEETS = {
    "test": ["code", "title", "description"],
    "scales": ["code", "title"],
    "questions": ["code", "text"],
    "answers": ["question_code", "answer_code", "text"],
    "scoring": ["question_code", "answer_code", "scale_code", "points"],
    "interpretations": ["scale_code", "min_score", "max_score", "text"],
}


class ExcelParserError(ValueError):
    pass


def parse_excel_test(path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ExcelParserError(
            "Для импорта Excel нужен пакет openpyxl. Установите зависимости из requirements.txt."
        ) from exc

    workbook = load_workbook(Path(path), data_only=True)
    rows_by_sheet = {
        sheet_name: _read_sheet(workbook, sheet_name, headers)
        for sheet_name, headers in EXPECTED_SHEETS.items()
    }

    test_row = _single_row(rows_by_sheet["test"], "test")
    questions_by_code = {
        row["code"]: {"text": row["text"], "answers": []}
        for row in rows_by_sheet["questions"]
    }

    scoring_by_answer = {}
    for row in rows_by_sheet["scoring"]:
        key = (row["question_code"], row["answer_code"])
        scoring_by_answer.setdefault(key, {})[row["scale_code"]] = int(row["points"] or 0)

    for row in rows_by_sheet["answers"]:
        question_code = row["question_code"]
        if question_code not in questions_by_code:
            raise ExcelParserError(f"Ответ ссылается на неизвестный вопрос: {question_code}")

        answer_code = row["answer_code"]
        questions_by_code[question_code]["answers"].append(
            {
                "code": answer_code,
                "text": row["text"],
                "scores": scoring_by_answer.get((question_code, answer_code), {}),
            }
        )

    return {
        "code": test_row["code"],
        "title": test_row["title"],
        "description": test_row["description"],
        "scales": [
            {"code": row["code"], "title": row["title"]}
            for row in rows_by_sheet["scales"]
        ],
        "questions": list(questions_by_code.values()),
        "interpretations": [
            {
                "scale": row["scale_code"] or "total",
                "min": int(row["min_score"]),
                "max": int(row["max_score"]),
                "text": row["text"],
            }
            for row in rows_by_sheet["interpretations"]
        ],
    }


def _read_sheet(workbook, sheet_name, required_headers):
    if sheet_name not in workbook.sheetnames:
        raise ExcelParserError(f"В Excel-файле нет листа {sheet_name}")

    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ExcelParserError(f"Лист {sheet_name} пуст")

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    missing_headers = [header for header in required_headers if header not in headers]
    if missing_headers:
        raise ExcelParserError(
            f"Лист {sheet_name}: нет колонок {', '.join(missing_headers)}"
        )

    result = []
    for row in rows[1:]:
        record = dict(zip(headers, row))
        if any(value is not None for value in record.values()):
            result.append({header: record.get(header) for header in headers})
    return result


def _single_row(rows, sheet_name):
    if len(rows) != 1:
        raise ExcelParserError(f"Лист {sheet_name} должен содержать ровно одну строку данных")
    return rows[0]
